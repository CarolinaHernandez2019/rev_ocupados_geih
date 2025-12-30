"""
Aplicación Streamlit para la Revisión de Ocupados - GEIH
Genera automáticamente los archivos de validación por posición ocupacional

Versión corregida que mantiene la estructura y lógica de los notebooks originales
"""

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

# =============================================================================
# CONFIGURACIÓN DE LA PÁGINA
# =============================================================================
st.set_page_config(
    page_title="Revisión Ocupados GEIH",
    page_icon="📊",
    layout="wide"
)

# =============================================================================
# ESTILOS Y CONSTANTES GLOBALES
# =============================================================================
ROJO = PatternFill('solid', fgColor='FF6B6B')
AMARILLO = PatternFill('solid', fgColor='FFE066')
VERDE = PatternFill('solid', fgColor='8FD14F')
AZUL = PatternFill('solid', fgColor='87CEEB')

ORDEN_RAMAS = [
    'No informa',
    'Agricultura, ganadería, caza, silvicultura y pesca',
    'Explotación de Minas y Canteras',
    'Comercio y reparación de vehículos',
    'Alojamiento y servicios de comida',
    'Industria manufacturera',
    'Suministro de agua y gestión de desechos',
    'Construcción',
    'Transporte y almacenamiento',
    'Información y comunicaciones',
    'Actividades financieras y de seguros',
    'Actividades Inmobiliarias',
    'Actividades profesionales, científicas, técnicas y servicios administrativos',
    'Administración pública y defensa, educación y atención de la salud',
    'Actividades artísticas, entretenimiento, recreación y otras actividades de servicios'
]

# =============================================================================
# DICCIONARIOS EMPLEADOS DEL GOBIERNO (P6430=2)
# =============================================================================

# Ramas donde NO debe haber empleados del gobierno
RAMAS_PROHIBIDAS_GOBIERNO = [
    'No informa',
    'Agricultura, ganadería, caza, silvicultura y pesca',
    'Explotación de Minas y Canteras',
    'Comercio y reparación de vehículos',
    'Construcción',
    'Alojamiento y servicios de comida'
]

# Ramas donde operan empresas mixtas/industriales del Estado
RAMAS_EMPRESAS_MIXTAS = [
    'Industria manufacturera',
    'Suministro de agua y gestión de desechos',
    'Transporte y almacenamiento',
    'Información y comunicaciones',
    'Actividades financieras y de seguros'
]

# Tipo de revisión por rama para gobierno
TIPO_REVISION_GOB = {
    'No informa': 1,
    'Agricultura, ganadería, caza, silvicultura y pesca': 1,
    'Explotación de Minas y Canteras': 2,
    'Comercio y reparación de vehículos': 1,
    'Construcción': 2,
    'Alojamiento y servicios de comida': 1,
    'Industria manufacturera': 2,
    'Suministro de agua y gestión de desechos': 2,
    'Transporte y almacenamiento': 2,
    'Información y comunicaciones': 2,
    'Actividades financieras y de seguros': 2,
    'Actividades Inmobiliarias': 1,
    'Actividades profesionales, científicas, técnicas y servicios administrativos': 2,
    'Administración pública y defensa, educación y atención de la salud': 0,
    'Actividades artísticas, entretenimiento, recreación y otras actividades de servicios': 2
}

# Entidades para cambio de rama
PALABRAS_RAMA_8412 = ['INVIAS', 'INSTITUTO NACIONAL DE VIAS', 'INVIR', 'ARCHIVO GENERAL', 'UNP', 
                      'UNIDAD NACIONAL DE PROTECCION', 'DIAN']
PALABRAS_RAMA_8414 = ['INSTITUTO COLOMBIANO AGROPECUARIO', ' ICA ', 'ICA-', 'AERONAUTICA CIVIL', 
                      'AEROCIVIL', 'DIMAR', 'ANI', 'AGENCIA NACIONAL DE INFRAESTRUCTURA',
                      'AGENCIA DE DESARROLLO RURAL', 'ADR', 'UNIDAD DE RESTITUCION', 'IGAC', 
                      'INSTITUTO GEOGRAFICO', 'SUPERINTENDENCIA', 'TRANSITO']
PALABRAS_RAMA_8413 = ['CORPORACION AUTONOMA', 'CAR ', 'CORPOAMAZONIA', 'CORTOLIMA', 'CORPOCALDAS',
                      'CORPOBOYACA', 'CORPONARIÑO', 'CRQ', 'CDA ', 'PARQUE NACIONAL',
                      'INDERVALLE', 'INDEPORTES', 'COLDEPORTES']
PALABRAS_RAMA_8424 = ['JUZGADO', 'FISCALIA', 'RAMA JUDICIAL', 'TRIBUNAL', 'PALACIO DE JUSTICIA',
                      'MEDICINA LEGAL', 'INPEC']
PALABRAS_RAMA_8415 = ['DEFENSORIA DEL PUEBLO', 'REGISTRADURIA', 'PERSONERIA']
PALABRAS_RAMA_8421 = ['MIGRACION COLOMBIA', 'CONSULADO', 'EMBAJADA', 'CANCILLERIA']

# Empresas con régimen laboral privado
EMPRESAS_REGIMEN_PRIVADO = ['ECOPETROL', 'CENIT']

# Entidades privadas - NO son gobierno
ENTIDADES_PRIVADAS_NO_GOBIERNO = ['CAMARA DE COMERCIO', 'FUNERARIA', 'NOTARIA']

# Empresas mixtas/industriales del Estado
EMPRESAS_MIXTAS = [
    # Energía
    'ISA ', 'ISAGEN', 'GECELCA', 'GENSA', 'CHEC', 'HIDROELECTRICA', 'ELECTRIFICADORA', 'CEELVA',
    # Servicios públicos
    'EMCALI', 'EPM', 'EMPRESAS PUBLICAS DE MEDELLIN', 'ACUEDUCTO', 'EAAB', 'ALCANTARILLADO',
    'EMPRESAS PUBLICAS DE', ' ESP', ' SA ESP', ' SAS ESP', 'EMPRESA DE SERVICIOS PUBLICOS',
    'SERVICIOS PUBLICOS DOMICILIARIOS', 'UNIDAD DE SERVICIOS PUBLICOS',
    # Agua
    'AGUAS DE ', 'AGUAS DEL ', 'AGUAS Y AGUAS', 'EMPAS', 'EMPOCALDAS', 'EMPOOBANDO',
    'EMPOCHIQUINQUIRA', 'ESSMAR', 'IBAL', 'SAAAB', 'ACUAVALLE', 'ACUAOCCIDENTE', 'PLANTA DE TRATAMIENTO',
    # Financieras
    'BANCO AGRARIO', 'FONDO NACIONAL DEL AHORRO', 'FNA ', 'COLPENSIONES', 'POSITIVA', 
    'FIDUPREVISORA', 'FINDETER', 'BANCOLDEX', 'FINAGRO', 'INFIBAGUE',
    # Manufactura estatal
    'LICORERA', 'INDUSTRIA LICORERA', 'INDUMIL', 'INDUSTRIA MILITAR', 'IMPRENTA NACIONAL', 'CIAC',
    # Transporte
    'METRO DE MEDELLIN', 'METRO DE BOGOTA', '472', 'SERVICIOS POSTALES', 'TERMINAL DE TRANSPORTE', 'SATENA',
    # Telecomunicaciones
    'ETB', 'EMPRESA DE TELECOMUNICACIONES', 'TELECARIBE', 'RTVC',
    # Otros
    'INNPULSA', 'SINCHI', 'LOTERIA', 'CORPOICA', 'AGROSAVIA', 'METROPARQUES', 
    'ARTESANIAS DE COLOMBIA', 'CISA'
]

# Cargos directivos
CARGOS_DIRECTIVOS_P6370 = ['PRESIDENTE', 'DIRECTOR', 'GERENTE', 'SUBGERENTE', 'VICEPRESIDENTE',
                           'SUBDIRECTOR', 'JEFE DE ', 'SECRETARIO GENERAL']
VALOR_DIRECTIVO_G_P6370S3 = 'Directores y gerentes'

# Entidades privadas en Adm. Pública
PALABRAS_PRIVADAS_ADM_PUBLICA = [
    'EPS ', 'SAVIA SALUD', 'ASMET SALUD', 'COMFACHOCO', 'NUEVA EPS', 'SANITAS', 'COOMEVA', 
    'SURA EPS', 'FAMISANAR', 'CLINICA ', 'HOSPITAL PRIVADO', 'FUNDACION ', 'HOGAR DE PASO', 
    'CENTRO DE BIENESTAR', 'COOPERATIVA', 'COOP ', 'ASOTRAINFA', 'GIMNASIO ',
    'S.A.S', ' SAS', ' LTDA', ' S.A.', ' S.A ', 'MI RED IPS'
]

# Contratantes gobierno
CONTRATANTES_GOBIERNO = [
    'SECRETARIA', 'MINISTERIO', 'ALCALDIA', 'GOBERNACION', 'DEPARTAMENTO', 'MUNICIPIO', 
    'GOBIERNO', 'ESTADO', 'ICBF', 'INSTITUTO COLOMBIANO DE BIENESTAR', 'BIENESTAR FAMILIAR',
    'SENA', 'EJERCITO', 'POLICIA', 'ARMADA', 'FUERZA AEREA', 'PROCURADURIA', 'CONTRALORIA', 
    'DEFENSORIA', 'DIAN', 'DANE', 'DNP', 'REGISTRADURIA', 'FISCALIA'
]

# Palabras que indican contratista
PALABRAS_CONTRATISTA = ['CONTRATISTA', 'PRESTACION DE SERVICIOS', 'PRESTACIÓN DE SERVICIOS',
                        'OPS', 'ORDEN DE PRESTACION', 'CONTRATO DE PRESTACION']


# =============================================================================
# DICCIONARIOS EMPLEADOS PARTICULARES (P6430=1)
# =============================================================================

# Universidades públicas
UNIVERSIDADES_PUBLICAS = [
    'UNIVERSIDAD NACIONAL', 'UNIVERSIDAD DE ANTIOQUIA', 'UNIVERSIDAD DEL VALLE', 
    'UNIVERSIDAD DE CARTAGENA', 'UNIVERSIDAD DEL CAUCA', 'UNIVERSIDAD DE CALDAS',
    'UNIVERSIDAD DE CORDOBA', 'UNIVERSIDAD DEL ATLANTICO', 'UNIVERSIDAD DEL MAGDALENA', 
    'UNIVERSIDAD DE NARIÑO', 'UNIVERSIDAD DEL TOLIMA', 'UNIVERSIDAD PEDAGOGICA',
    'UNIVERSIDAD TECNOLOGICA DE PEREIRA', 'UTP ', 'UNIVERSIDAD SURCOLOMBIANA',
    'UNIVERSIDAD DE PAMPLONA', 'UNIVERSIDAD DE LOS LLANOS', 'UNIVERSIDAD DE LA GUAJIRA',
    'UNIVERSIDAD FRANCISCO DE PAULA', 'UFPS', 'UNIVERSIDAD DISTRITAL'
]

# Entidades del gobierno
ENTIDADES_GOBIERNO = [
    'MINISTERIO DE', 'MINISTERIO DEL', 'DEPARTAMENTO ADMINISTRATIVO NACIONAL DE ESTADISTICA',
    'DEPARTAMENTO NACIONAL DE PLANEACION', 'DIRECCION DE IMPUESTOS Y ADUANAS',
    'INSTITUTO COLOMBIANO', 'ICBF', ' SENA', 'INVIAS', 'INPEC', 'ICFES',
    'FISCALIA', 'PROCURADURIA', 'CONTRALORIA', 'DEFENSORIA', 'REGISTRADURIA',
    'POLICIA NACIONAL', 'EJERCITO NACIONAL', 'ARMADA NACIONAL', 'FUERZA AEREA',
    'ALCALDIA', 'GOBERNACION', 'SECRETARIA DE', 'SECRETARIA DISTRITAL',
    'CONCEJO', 'ASAMBLEA', 'CONGRESO', 'SENADO', 'CAMARA DE REPRESENTANTES',
    'HOSPITAL DEPARTAMENTAL', 'HOSPITAL MUNICIPAL', 'E.S.E', 'ESE ', ' ESE',
    'PERSONERIA', 'JUZGADO', 'TRIBUNAL'
]

# Instituciones educativas públicas
INSTITUCIONES_EDUCATIVAS_PUBLICAS = ['INSTITUCION EDUCATIVA ', 'I.E. ', 'I.E.D.',
                                     'COLEGIO DISTRITAL', 'COLEGIO DEPARTAMENTAL', 'COLEGIO MUNICIPAL']
INDICADORES_IE_PRIVADA = ['CRISTIANA', 'CRISTIANO', 'EVANGELICA', 'EVANGELICO',
                          'CATOLICA', 'CATOLICO', 'ADVENTISTA', 'BAUTISTA',
                          'BILINGUE', 'CAMPESTRE', 'INTERNACIONAL', 'PRIVAD']

# Empresas privadas
EMPRESAS_PRIVADAS = ['S.A.S', ' SAS', 'LTDA', 'S.A.', ' SA ', 'CLINICA ', 'EPS ', 'IPS ', 
                     'COLSANITAS', 'SANITAS', 'COOMEVA', 'SURA ', 'NUEVA EPS', 'COMPENSAR',
                     'NOTARIA ', 'FUNERARIA']

# Palabras para jornalero
PALABRAS_PRODUCCION_DIRECTA = ['ORDEÑ', 'ORDENA', 'SEMBRAR', 'SIEMBRA', 'PLANTAR',
                               'RECOLECT', 'COSECH', 'CORTAR CAÑA', 'CORTERO',
                               'FUMIG', 'ABON', 'FERTILIZ', 'DESHIERB', 'DESYERB', 
                               'GUADAÑ', 'CHAPEAR', 'ROZAR', 'JORNALERO', 'PEON',
                               'ALIMENTAR GANADO', 'ARREAR', 'PASTOREAR']

PALABRAS_SUPERVISION = ['DIRIGIR', 'DIRIGE', 'DIRECCION', 'ADMINISTR', 'GERENTE', 'GERENCIA',
                        'COORDINAR', 'COORDINADOR', 'PLANEAR', 'PLANIFICA', 'PLANEACION',
                        'SUPERVISAR', 'SUPERVISOR', 'MAYORDOMO', 'CAPATAZ', 'ENCARGADO DE FINCA']

# Palabras para empleado doméstico
PALABRAS_DOMESTICO = ['EMPLEADA DOMESTICA', 'EMPLEADO DOMESTICO', 'SERVICIO DOMESTICO',
                      'ASEO EN CASA', 'HOGAR ', 'OFICIO DE LA CASA', 'LABORES DOMESTICAS',
                      'NIÑERA', 'CUIDAR NIÑOS', 'CUIDADO DE NIÑOS']


# =============================================================================
# DICCIONARIOS TRABAJADOR FAMILIAR (P6430=6)
# =============================================================================

ENTIDADES_NO_FAMILIARES = [
    # Entidades religiosas
    'IGLESIA', 'PARROQUIA', 'TEMPLO', 'CAPILLA', 'CATEDRAL', 'DIOCESIS', 'ARQUIDIOCESIS', 
    'CONGREGACION', 'COMUNIDAD RELIGIOSA',
    # Entidades públicas
    'ALCALDIA', 'GOBERNACION', 'MINISTERIO', 'SECRETARIA DE', 'INSTITUTO COLOMBIANO', 
    'ICBF', 'SENA', 'POLICIA', 'EJERCITO', 'FISCALIA', 'PROCURADURIA', 'CONTRALORIA', 
    'JUZGADO', 'TRIBUNAL', 'UNIVERSIDAD NACIONAL', 'UNIVERSIDAD DE ANTIOQUIA', 
    'UNIVERSIDAD DEL VALLE', 'INSTITUCION EDUCATIVA', 'I.E.', 'E.S.E.', 'HOSPITAL DEPARTAMENTAL',
    # Empresas formales
    'S.A.S', 'SAS', 'S.A', 'LTDA', 'LIMITADA', 'E.S.P', 'ESP',
    'BANCO', 'ALMACEN', 'SUPERMERCADO', 'EXITO', 'JUMBO', 'CARULLA', 'OLIMPICA',
    'FUNDACION', 'CORPORACION', 'COOPERATIVA', 'ONG'
]

CARGOS_DECISION = [
    'DUEÑO', 'DUEÑA', 'PROPIETARIO', 'PROPIETARIA', 'SOCIO', 'SOCIA', 'ACCIONISTA',
    'GERENTE', 'DIRECTOR', 'DIRECTORA', 'ADMINISTRADOR GENERAL', 'ADMINISTRADORA GENERAL',
    'REPRESENTANTE LEGAL', 'MI NEGOCIO', 'MI EMPRESA', 'NEGOCIO PROPIO', 'EMPRESA PROPIA',
    'SU PROPIO NEGOCIO'
]

INDICADORES_FAMILIAR = [
    'TIENDA ', 'MISCELANEA', 'PAPELERIA', 'PANADERIA', 'FERRETERIA', 'DROGUERIA', 
    'PELUQUERIA', 'BARBERIA', 'RESTAURANTE ', 'CAFETERIA', 'FRUTERIA', 'CARNICERIA',
    'TALLER ', 'SASTRERIA', 'MODISTERIA', 'LAVADERO', 'FINCA ', 'PARCELA', 'HACIENDA',
    'DONDE ', 'DE ', 'LA ', 'EL ', 'LOS ', 'LAS '
]


# =============================================================================
# DICCIONARIOS OTRO CUÁL (P6430=8)
# =============================================================================

PALABRAS_CUENTA_PROPIA = [
    'CONTRATISTA', 'PRESTACION DE SERVICIOS', 'PRESTACIÓN DE SERVICIOS',
    'CONTRATO DE PRESTACION', 'CONTRATO DE PRESTACIÓN', 'PRESTA SERVICIOS',
    'INDEPENDIENTE', 'FREELANCE', 'FREELANCER', 'POR SU CUENTA', 'TRABAJO INDEPENDIENTE'
]

PALABRAS_PATRON = ['SOCIO', 'SOCIA', 'DUEÑO', 'DUEÑA', 'PROPIETARIO', 'PROPIETARIA', 
                   'ACCIONISTA', 'EMPRESARIO']

PALABRAS_OTRO_VALIDO = [
    'SUBCONTRATADO', 'SUBCONTRATADA', 'CONTRATADO POR UN ASALARIADO', 'CONTRATADA POR UN ASALARIADO',
    'CONTRATADO POR TRABAJADOR', 'CONTRATADA POR TRABAJADOR', 'EMPLEADO DE UN INDEPENDIENTE',
    'EMPLEADA DE UN INDEPENDIENTE', 'TRABAJA PARA UN ASALARIADO', 'TRABAJA PARA UNA ASALARIADA',
    'CONTRATADO POR OTRA PERSONA', 'MADRE COMUNITARIA', 'AYUDANTE DE MADRE', 'OTRO PAIS', 
    'OTRO PAÍS', 'TRABAJA EN OTRO', 'HIJO DEL MAYORDOMO', 'HIJA DEL MAYORDOMO'
]


# =============================================================================
# FUNCIONES DE CLASIFICACIÓN
# =============================================================================

def es_directivo(row):
    """Verifica si la persona ocupa un cargo directivo."""
    g_p6370s3 = str(row.get('g_p6370s3', '')).strip() if pd.notna(row.get('g_p6370s3')) else ''
    if VALOR_DIRECTIVO_G_P6370S3.lower() in g_p6370s3.lower():
        return True
    p6370 = str(row.get('p6370', '')).upper() if pd.notna(row.get('p6370')) else ''
    for cargo in CARGOS_DIRECTIVOS_P6370:
        if cargo in p6370:
            return True
    return False


def clasificar_empleado_gobierno(row):
    """
    Clasifica empleados del gobierno (P6430=2).
    Retorna dict con: tipo_revision, pos_corregida, rama_corregida, observacion
    """
    rama = str(row.get('g_p6390s2', '')) if pd.notna(row.get('g_p6390s2')) else ''
    empresa = str(row.get('p6380', '')).upper() if pd.notna(row.get('p6380')) else ''
    oficio = str(row.get('p6370', '')).upper() if pd.notna(row.get('p6370')) else ''
    p6400 = row.get('p6400', None)
    
    resultado = {'tipo_revision': 0, 'pos_corregida': None, 'rama_corregida': None, 'observacion': ''}
    
    # 1. Empresas con régimen laboral privado (Ecopetrol)
    if any(emp in empresa for emp in EMPRESAS_REGIMEN_PRIVADO):
        resultado['tipo_revision'] = 1
        resultado['pos_corregida'] = 1
        resultado['observacion'] = 'CAMBIAR → Pos 1: Empresa con régimen laboral privado (Ley 1118/2006)'
        return resultado
    
    # 2. Entidades privadas (Cámara de Comercio, Notarías)
    if any(ent in empresa for ent in ENTIDADES_PRIVADAS_NO_GOBIERNO):
        resultado['tipo_revision'] = 1
        resultado['pos_corregida'] = 1
        resultado['observacion'] = 'CAMBIAR → Pos 1: Entidad privada, no es gobierno'
        return resultado
    
    # 3. Revisar por tipo de rama
    tipo_rama = TIPO_REVISION_GOB.get(rama, 0)
    
    # Rama prohibida (tipo 1)
    if tipo_rama == 1:
        # Verificar si es cambio de rama en vez de posición
        for patron in PALABRAS_RAMA_8412:
            if patron in empresa:
                resultado['tipo_revision'] = 2
                resultado['rama_corregida'] = '8412'
                resultado['observacion'] = 'CAMBIAR RAMA → 8412: Actividades ejecutivas administración pública'
                return resultado
        for patron in PALABRAS_RAMA_8414:
            if patron in empresa:
                resultado['tipo_revision'] = 2
                resultado['rama_corregida'] = '8414'
                resultado['observacion'] = 'CAMBIAR RAMA → 8414: Actividades reguladoras'
                return resultado
        for patron in PALABRAS_RAMA_8413:
            if patron in empresa:
                resultado['tipo_revision'] = 2
                resultado['rama_corregida'] = '8413'
                resultado['observacion'] = 'CAMBIAR RAMA → 8413: Programas bienestar/medio ambiente'
                return resultado
        
        resultado['tipo_revision'] = 1
        resultado['pos_corregida'] = 1
        resultado['observacion'] = 'CAMBIAR → Pos 1: Rama prohibida para empleado gobierno'
        return resultado
    
    # Empresas mixtas (tipo 2)
    if tipo_rama == 2 or any(emp in empresa for emp in EMPRESAS_MIXTAS):
        if es_directivo(row):
            resultado['tipo_revision'] = 4
            resultado['observacion'] = 'REVISAR: Directivo en empresa mixta (verificar si es EICE)'
        else:
            resultado['tipo_revision'] = 1
            resultado['pos_corregida'] = 1
            resultado['observacion'] = 'CAMBIAR → Pos 1: No directivo en empresa mixta'
        return resultado
    
    # Administración pública (tipo 0 o 3)
    if rama == 'Administración pública y defensa, educación y atención de la salud':
        # Verificar si es entidad privada
        if any(p in empresa for p in PALABRAS_PRIVADAS_ADM_PUBLICA):
            resultado['tipo_revision'] = 1
            resultado['pos_corregida'] = 1
            resultado['observacion'] = 'CAMBIAR → Pos 1: Entidad privada en rama Adm. Pública'
            return resultado
        
        # Verificar contratistas
        if any(c in oficio or c in empresa for c in PALABRAS_CONTRATISTA):
            resultado['tipo_revision'] = 1
            resultado['pos_corregida'] = 5
            resultado['observacion'] = 'CAMBIAR → Pos 5: Contratista/Prestador de servicios'
            return resultado
        
        # Verificar intermediación
        try:
            if pd.notna(p6400) and int(p6400) == 2:
                resultado['tipo_revision'] = 4
                resultado['observacion'] = 'REVISAR: Trabaja por intermediación (P6400=2)'
                return resultado
        except:
            pass
        
        resultado['observacion'] = 'OK'
        return resultado
    
    # Otras ramas
    resultado['tipo_revision'] = 4
    resultado['observacion'] = 'REVISAR: Verificar si la entidad es pública'
    return resultado


def clasificar_empleado_particular(row):
    """
    Clasifica empleado particular (P6430=1).
    Detecta casos que deberían ser: Gobierno (2), Doméstico (3), Jornalero (7)
    """
    rama = str(row.get('g_p6390s2', '')).upper() if pd.notna(row.get('g_p6390s2')) else ''
    empresa = str(row.get('p6380', '')).upper() if pd.notna(row.get('p6380')) else ''
    oficio = str(row.get('p6370', '')).upper() if pd.notna(row.get('p6370')) else ''
    
    resultado = {'tipo_revision': 0, 'pos_corregida': None, 'observacion': ''}
    
    # 1. Posible empleado gobierno - Universidades públicas
    for uni in UNIVERSIDADES_PUBLICAS:
        if uni in empresa:
            resultado['tipo_revision'] = 1
            resultado['pos_corregida'] = 2
            resultado['observacion'] = 'REVISAR → Pos 2: Universidad pública'
            return resultado
    
    # 2. Posible empleado gobierno - Entidades del gobierno
    for ent in ENTIDADES_GOBIERNO:
        if ent in empresa:
            # Excluir si tiene indicadores de privado
            if not any(p in empresa for p in ['CLINICA ', ' SAS', 'S.A.S', 'LTDA']):
                resultado['tipo_revision'] = 1
                resultado['pos_corregida'] = 2
                resultado['observacion'] = 'REVISAR → Pos 2: Posible entidad del gobierno'
                return resultado
    
    # 3. Posible empleado gobierno - Instituciones educativas públicas
    for ie in INSTITUCIONES_EDUCATIVAS_PUBLICAS:
        if ie in empresa:
            if not any(priv in empresa for priv in INDICADORES_IE_PRIVADA):
                resultado['tipo_revision'] = 1
                resultado['pos_corregida'] = 2
                resultado['observacion'] = 'REVISAR → Pos 2: Institución educativa pública'
                return resultado
    
    # 4. Posible empleado doméstico
    if any(d in oficio or d in empresa for d in PALABRAS_DOMESTICO):
        resultado['tipo_revision'] = 2
        resultado['pos_corregida'] = 3
        resultado['observacion'] = 'REVISAR → Pos 3: Posible empleado doméstico'
        return resultado
    
    # 5. Posible jornalero (solo en Agricultura)
    if 'AGRICULTURA' in rama:
        # Verificar si es supervisión (NO es jornalero)
        if any(s in oficio for s in PALABRAS_SUPERVISION):
            resultado['observacion'] = 'OK: Supervisión en agricultura'
            return resultado
        
        # Verificar si es producción directa (SÍ es jornalero)
        if any(p in oficio for p in PALABRAS_PRODUCCION_DIRECTA):
            resultado['tipo_revision'] = 3
            resultado['pos_corregida'] = 7
            resultado['observacion'] = 'REVISAR → Pos 7: Posible jornalero (producción directa)'
            return resultado
    
    resultado['observacion'] = 'OK'
    return resultado


def clasificar_trabajador_familiar(row):
    """
    Clasifica trabajador familiar sin remuneración (P6430=6).
    """
    empresa = str(row.get('p6380', '')).upper() if pd.notna(row.get('p6380')) else ''
    oficio = str(row.get('p6370', '')).upper() if pd.notna(row.get('p6370')) else ''
    p3069 = row.get('p3069', None)
    
    resultado = {'tipo_revision': 0, 'pos_corregida': None, 'observacion': ''}
    
    # 1. Trabaja solo (P3069=1) - No puede ser trabajador familiar
    try:
        if pd.notna(p3069) and int(p3069) == 1:
            resultado['tipo_revision'] = 1
            resultado['observacion'] = 'DETALLAR: Trabaja solo (P3069=1) - No puede ser familiar'
            return resultado
    except:
        pass
    
    # 2. Entidad no familiar
    if any(e in empresa for e in ENTIDADES_NO_FAMILIARES):
        resultado['tipo_revision'] = 2
        resultado['observacion'] = 'DETALLAR: Entidad no familiar (iglesia, empresa formal, etc.)'
        return resultado
    
    # 3. Cargo de decisión → posible cuenta propia
    texto = f"{oficio} {empresa}"
    if any(c in texto for c in CARGOS_DECISION):
        resultado['tipo_revision'] = 3
        resultado['pos_corregida'] = 5
        resultado['observacion'] = 'DETALLAR → Pos 5: Cargo decisión (dueño/socio/gerente)'
        return resultado
    
    # 4. Verificar si parece empresa familiar (OK)
    if any(i in empresa for i in INDICADORES_FAMILIAR):
        resultado['observacion'] = 'OK: Parece empresa familiar'
    else:
        resultado['tipo_revision'] = 4
        resultado['observacion'] = 'REVISAR: Verificar si es empresa familiar'
    
    return resultado


def clasificar_otro_cual(row):
    """
    Clasifica 'Otro, ¿cuál?' (P6430=8).
    """
    oficio = str(row.get('p6370', '')).upper() if pd.notna(row.get('p6370')) else ''
    otro_cual = str(row.get('p6430s1', '')).upper() if pd.notna(row.get('p6430s1')) else ''
    empresa = str(row.get('p6380', '')).upper() if pd.notna(row.get('p6380')) else ''
    p3069 = row.get('p3069', None)
    
    texto = f"{oficio} {otro_cual} {empresa}"
    resultado = {'tipo_revision': 0, 'pos_corregida': None, 'observacion': ''}
    
    # 1. Contratista/Independiente → Cuenta propia
    if any(p in texto for p in PALABRAS_CUENTA_PROPIA):
        resultado['tipo_revision'] = 1
        resultado['pos_corregida'] = 5
        resultado['observacion'] = 'CAMBIAR → Pos 5: Contratista/Independiente es cuenta propia'
        return resultado
    
    # 2. Socio/Dueño → Patrón o Cuenta propia
    if any(p in texto for p in PALABRAS_PATRON):
        tiene_empleados = False
        try:
            if pd.notna(p3069) and int(p3069) > 1:
                tiene_empleados = True
        except:
            pass
        
        if tiene_empleados:
            resultado['tipo_revision'] = 2
            resultado['pos_corregida'] = 4
            resultado['observacion'] = 'CAMBIAR → Pos 4: Socio/Dueño con empleados es patrón'
        else:
            resultado['tipo_revision'] = 1
            resultado['pos_corregida'] = 5
            resultado['observacion'] = 'CAMBIAR → Pos 5: Socio/Dueño sin empleados es cuenta propia'
        return resultado
    
    # 3. Caso válido de "Otro"
    if any(v in texto for v in PALABRAS_OTRO_VALIDO):
        resultado['observacion'] = 'OK: Caso válido de "Otro"'
        return resultado
    
    # 4. Sin clasificar - revisar descripción
    if len(otro_cual.strip()) > 3:
        resultado['tipo_revision'] = 3
        resultado['observacion'] = f'DETALLAR: Verificar descripción "{otro_cual[:50]}"'
    else:
        resultado['tipo_revision'] = 3
        resultado['observacion'] = 'DETALLAR: Sin descripción clara en P6430S1'
    
    return resultado


# =============================================================================
# FUNCIONES PARA GENERAR EXCEL
# =============================================================================

def generar_excel_gobierno(df_gob):
    """Genera Excel para empleados del gobierno con estructura del notebook original."""
    if len(df_gob) == 0:
        return None
    
    # Aplicar clasificación
    clasificaciones = df_gob.apply(clasificar_empleado_gobierno, axis=1, result_type='expand')
    df_gob = df_gob.copy()
    df_gob['tipo_revision'] = clasificaciones['tipo_revision']
    df_gob['pos_corregida'] = clasificaciones['pos_corregida']
    df_gob['rama_corregida'] = clasificaciones['rama_corregida']
    df_gob['observacion'] = clasificaciones['observacion']
    
    # Crear resumen por rama
    resumen = df_gob.groupby('g_p6390s2').agg(
        Casos=('directorio', 'count'),
        Cambiar_Pos=('pos_corregida', lambda x: x.notna().sum()),
        Cambiar_Rama=('rama_corregida', lambda x: x.notna().sum())
    ).reset_index()
    resumen = resumen.rename(columns={'g_p6390s2': 'RAMA DE ACTIVIDAD ECONÓMICA'})
    
    # Contar casos a revisar (tipo 4)
    revisar_counts = df_gob[df_gob['tipo_revision'] == 4].groupby('g_p6390s2').size()
    resumen = resumen.merge(
        revisar_counts.reset_index().rename(columns={'g_p6390s2': 'RAMA DE ACTIVIDAD ECONÓMICA', 0: 'Revisar'}),
        on='RAMA DE ACTIVIDAD ECONÓMICA', how='left'
    ).fillna(0)
    resumen['Revisar'] = resumen['Revisar'].astype(int)
    
    # Ordenar por ORDEN_RAMAS
    todas_ramas = pd.DataFrame({'RAMA DE ACTIVIDAD ECONÓMICA': ORDEN_RAMAS})
    resumen = todas_ramas.merge(resumen, on='RAMA DE ACTIVIDAD ECONÓMICA', how='left').fillna(0)
    for col in ['Casos', 'Cambiar_Pos', 'Cambiar_Rama', 'Revisar']:
        resumen[col] = resumen[col].astype(int)
    
    # Agregar total
    total = pd.DataFrame([{
        'RAMA DE ACTIVIDAD ECONÓMICA': 'TOTAL',
        'Casos': resumen['Casos'].sum(),
        'Cambiar_Pos': resumen['Cambiar_Pos'].sum(),
        'Cambiar_Rama': resumen['Cambiar_Rama'].sum(),
        'Revisar': resumen['Revisar'].sum()
    }])
    resumen = pd.concat([resumen, total], ignore_index=True)
    
    # Crear cuadro de inconsistencias
    inconsistencias = df_gob[df_gob['tipo_revision'] > 0].copy()
    
    # Crear Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # HOJA 1: RESUMEN CON SEMÁFORO
        resumen.to_excel(writer, sheet_name='Resumen', index=False, startrow=4)
        ws = writer.sheets['Resumen']
        
        ws['A1'] = 'DISTRIBUCIÓN DE EMPLEADOS DEL GOBIERNO (P6430=2) POR RAMA DE ACTIVIDAD'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = 'NOTA: Cambiar_Pos = Cambio de posición ocupacional | Cambiar_Rama = Cambio de rama de actividad'
        ws['A2'].font = Font(italic=True, size=10)
        ws['A3'] = 'SEMÁFORO:'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = '🔴 ROJO = Casos a cambiar'
        ws['B3'].fill = ROJO
        ws['C3'] = '🟡 AMARILLO = Cambio rama'
        ws['C3'].fill = AMARILLO
        ws['D3'] = '🔵 AZUL = Revisar'
        ws['D3'].fill = AZUL
        ws['E3'] = '🟢 VERDE = OK'
        ws['E3'].fill = VERDE
        
        # Aplicar colores semáforo
        for i, row in resumen.iterrows():
            fila_excel = i + 6
            if row['RAMA DE ACTIVIDAD ECONÓMICA'] != 'TOTAL':
                if row['Cambiar_Pos'] > 0:
                    color = ROJO
                elif row['Cambiar_Rama'] > 0:
                    color = AMARILLO
                elif row['Revisar'] > 0:
                    color = AZUL
                else:
                    color = VERDE
                for col in range(1, 6):
                    ws.cell(row=fila_excel, column=col).fill = color
        
        # Formato fila total
        fila_total = len(resumen) + 5
        for col in range(1, 6):
            cell = ws.cell(row=fila_total, column=col)
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
        
        ws.column_dimensions['A'].width = 70
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
        
        # HOJA 2: INCONSISTENCIAS (cuadro resumen)
        if len(inconsistencias) > 0:
            # Crear cuadro pivote de inconsistencias
            cuadro_inc = inconsistencias.groupby('g_p6390s2').agg(
                Cambiar_Pos=('pos_corregida', lambda x: x.notna().sum()),
                Cambiar_Rama=('rama_corregida', lambda x: x.notna().sum()),
                Revisar=('tipo_revision', lambda x: (x == 4).sum())
            ).reset_index()
            cuadro_inc = cuadro_inc.rename(columns={'g_p6390s2': 'RAMA'})
            cuadro_inc['TOTAL'] = cuadro_inc['Cambiar_Pos'] + cuadro_inc['Cambiar_Rama'] + cuadro_inc['Revisar']
            cuadro_inc = cuadro_inc[cuadro_inc['TOTAL'] > 0]
            
            # Agregar fila total
            total_inc = pd.DataFrame([{
                'RAMA': 'TOTAL',
                'Cambiar_Pos': cuadro_inc['Cambiar_Pos'].sum(),
                'Cambiar_Rama': cuadro_inc['Cambiar_Rama'].sum(),
                'Revisar': cuadro_inc['Revisar'].sum(),
                'TOTAL': cuadro_inc['TOTAL'].sum()
            }])
            cuadro_inc = pd.concat([cuadro_inc, total_inc], ignore_index=True)
            
            cuadro_inc.to_excel(writer, sheet_name='Inconsistencias', index=False, startrow=2)
            ws2 = writer.sheets['Inconsistencias']
            ws2['A1'] = 'RESUMEN DE CASOS A REVISAR'
            ws2['A1'].font = Font(bold=True, size=12)
            ws2.column_dimensions['A'].width = 70
        
        # HOJA 3: CASOS PARA REVISIÓN (columnas acotadas)
        cols_revision = ['directorio', 'secuencia_p', 'orden', 'municipio', 
                        'p6370', 'p6380', 'g_p6390s2', 'p6400',
                        'tipo_revision', 'pos_corregida', 'rama_corregida', 'observacion']
        cols_disponibles = [c for c in cols_revision if c in df_gob.columns]
        casos_rev = df_gob[df_gob['tipo_revision'] > 0][cols_disponibles].copy()
        casos_rev.to_excel(writer, sheet_name='Casos_Revision', index=False)
        
        # HOJA 4: TODOS LOS CASOS (base completa)
        df_gob.to_excel(writer, sheet_name='Casos_Completo', index=False)
    
    output.seek(0)
    return output


def generar_excel_particular(df_part):
    """Genera Excel para empleados particulares."""
    if len(df_part) == 0:
        return None
    
    # Aplicar clasificación
    clasificaciones = df_part.apply(clasificar_empleado_particular, axis=1, result_type='expand')
    df_part = df_part.copy()
    df_part['tipo_revision'] = clasificaciones['tipo_revision']
    df_part['pos_corregida'] = clasificaciones['pos_corregida']
    df_part['observacion'] = clasificaciones['observacion']
    
    # Crear resumen
    resumen = df_part.groupby('g_p6390s2').agg(
        Casos=('directorio', 'count'),
        Revisar_Gobierno=('tipo_revision', lambda x: (x == 1).sum()),
        Revisar_Domestico=('tipo_revision', lambda x: (x == 2).sum()),
        Revisar_Jornalero=('tipo_revision', lambda x: (x == 3).sum())
    ).reset_index()
    resumen = resumen.rename(columns={'g_p6390s2': 'RAMA DE ACTIVIDAD ECONÓMICA'})
    
    # Ordenar y agregar total
    todas_ramas = pd.DataFrame({'RAMA DE ACTIVIDAD ECONÓMICA': ORDEN_RAMAS})
    resumen = todas_ramas.merge(resumen, on='RAMA DE ACTIVIDAD ECONÓMICA', how='left').fillna(0)
    for col in ['Casos', 'Revisar_Gobierno', 'Revisar_Domestico', 'Revisar_Jornalero']:
        resumen[col] = resumen[col].astype(int)
    
    total = pd.DataFrame([{
        'RAMA DE ACTIVIDAD ECONÓMICA': 'TOTAL',
        'Casos': resumen['Casos'].sum(),
        'Revisar_Gobierno': resumen['Revisar_Gobierno'].sum(),
        'Revisar_Domestico': resumen['Revisar_Domestico'].sum(),
        'Revisar_Jornalero': resumen['Revisar_Jornalero'].sum()
    }])
    resumen = pd.concat([resumen, total], ignore_index=True)
    
    # Crear Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # HOJA 1: RESUMEN
        resumen.to_excel(writer, sheet_name='Resumen', index=False, startrow=4)
        ws = writer.sheets['Resumen']
        
        ws['A1'] = 'DISTRIBUCIÓN DE EMPLEADOS PARTICULARES (P6430=1) POR RAMA DE ACTIVIDAD'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = 'Detecta posibles cambios a: Gobierno (Pos 2), Doméstico (Pos 3), Jornalero (Pos 7)'
        ws['A2'].font = Font(italic=True, size=10)
        ws['A3'] = 'SEMÁFORO:'
        ws['B3'] = '🟡 AMARILLO = Casos a revisar'
        ws['B3'].fill = AMARILLO
        ws['C3'] = '🟢 VERDE = OK'
        ws['C3'].fill = VERDE
        
        # Aplicar colores
        for i, row in resumen.iterrows():
            fila_excel = i + 6
            if row['RAMA DE ACTIVIDAD ECONÓMICA'] != 'TOTAL':
                total_revisar = row['Revisar_Gobierno'] + row['Revisar_Domestico'] + row['Revisar_Jornalero']
                color = AMARILLO if total_revisar > 0 else VERDE
                for col in range(1, 6):
                    ws.cell(row=fila_excel, column=col).fill = color
        
        ws.column_dimensions['A'].width = 70
        
        # HOJA 2: CASOS REVISIÓN
        cols_revision = ['directorio', 'secuencia_p', 'orden', 'municipio',
                        'p6370', 'p6380', 'g_p6390s2', 'p6400',
                        'tipo_revision', 'pos_corregida', 'observacion']
        cols_disponibles = [c for c in cols_revision if c in df_part.columns]
        casos_rev = df_part[df_part['tipo_revision'] > 0][cols_disponibles].copy()
        casos_rev.to_excel(writer, sheet_name='Casos_Revision', index=False)
        
        # HOJA 3: TODOS LOS CASOS
        df_part.to_excel(writer, sheet_name='Casos_Completo', index=False)
    
    output.seek(0)
    return output


def generar_excel_familiar(df_fam):
    """Genera Excel para trabajadores familiares."""
    if len(df_fam) == 0:
        return None
    
    # Aplicar clasificación
    clasificaciones = df_fam.apply(clasificar_trabajador_familiar, axis=1, result_type='expand')
    df_fam = df_fam.copy()
    df_fam['tipo_revision'] = clasificaciones['tipo_revision']
    df_fam['pos_corregida'] = clasificaciones['pos_corregida']
    df_fam['observacion'] = clasificaciones['observacion']
    
    # Crear resumen
    resumen = df_fam.groupby('g_p6390s2').agg(
        Casos=('directorio', 'count'),
        Detallar=('tipo_revision', lambda x: (x.isin([1, 2, 3])).sum()),
        Revisar=('tipo_revision', lambda x: (x == 4).sum())
    ).reset_index()
    resumen = resumen.rename(columns={'g_p6390s2': 'RAMA DE ACTIVIDAD ECONÓMICA'})
    
    # Ordenar y agregar total
    todas_ramas = pd.DataFrame({'RAMA DE ACTIVIDAD ECONÓMICA': ORDEN_RAMAS})
    resumen = todas_ramas.merge(resumen, on='RAMA DE ACTIVIDAD ECONÓMICA', how='left').fillna(0)
    for col in ['Casos', 'Detallar', 'Revisar']:
        resumen[col] = resumen[col].astype(int)
    
    total = pd.DataFrame([{
        'RAMA DE ACTIVIDAD ECONÓMICA': 'TOTAL',
        'Casos': resumen['Casos'].sum(),
        'Detallar': resumen['Detallar'].sum(),
        'Revisar': resumen['Revisar'].sum()
    }])
    resumen = pd.concat([resumen, total], ignore_index=True)
    
    # Crear Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # HOJA 1: RESUMEN
        resumen.to_excel(writer, sheet_name='Resumen', index=False, startrow=4)
        ws = writer.sheets['Resumen']
        
        ws['A1'] = 'DISTRIBUCIÓN DE TRABAJADORES FAMILIARES SIN REMUNERACIÓN (P6430=6) POR RAMA'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = 'NOTA: Los casos a DETALLAR deben devolverse a campo (flujo diferente al de asalariados)'
        ws['A2'].font = Font(italic=True, size=10)
        ws['A3'] = 'SEMÁFORO:'
        ws['B3'] = '🔴 ROJO = Casos a detallar'
        ws['B3'].fill = ROJO
        ws['C3'] = '🟡 AMARILLO = Casos a revisar'
        ws['C3'].fill = AMARILLO
        ws['D3'] = '🟢 VERDE = OK'
        ws['D3'].fill = VERDE
        
        # Aplicar colores
        for i, row in resumen.iterrows():
            fila_excel = i + 6
            if row['RAMA DE ACTIVIDAD ECONÓMICA'] != 'TOTAL':
                if row['Detallar'] > 0:
                    color = ROJO
                elif row['Revisar'] > 0:
                    color = AMARILLO
                else:
                    color = VERDE
                for col in range(1, 5):
                    ws.cell(row=fila_excel, column=col).fill = color
        
        ws.column_dimensions['A'].width = 70
        
        # HOJA 2: CUADRO INCONSISTENCIAS
        inconsistencias = df_fam[df_fam['tipo_revision'] > 0].copy()
        if len(inconsistencias) > 0:
            def categorizar(row):
                tipos = {1: 'TRABAJA_SOLO', 2: 'ENTIDAD_NO_FAMILIAR', 
                        3: 'CARGO_DECISION', 4: 'REVISAR'}
                return tipos.get(row['tipo_revision'], 'OTRO')
            
            inconsistencias['categoria'] = inconsistencias.apply(categorizar, axis=1)
            cuadro_inc = inconsistencias.groupby(['g_p6390s2', 'categoria']).size().unstack(fill_value=0)
            cuadro_inc = cuadro_inc.reset_index().rename(columns={'g_p6390s2': 'RAMA'})
            
            for col in ['TRABAJA_SOLO', 'ENTIDAD_NO_FAMILIAR', 'CARGO_DECISION', 'REVISAR']:
                if col not in cuadro_inc.columns:
                    cuadro_inc[col] = 0
            
            cols_orden = ['RAMA', 'TRABAJA_SOLO', 'ENTIDAD_NO_FAMILIAR', 'CARGO_DECISION', 'REVISAR']
            cuadro_inc = cuadro_inc[cols_orden]
            cuadro_inc['TOTAL'] = cuadro_inc[['TRABAJA_SOLO', 'ENTIDAD_NO_FAMILIAR', 'CARGO_DECISION', 'REVISAR']].sum(axis=1)
            
            # Agregar total
            total_inc = pd.DataFrame([{
                'RAMA': 'TOTAL',
                'TRABAJA_SOLO': cuadro_inc['TRABAJA_SOLO'].sum(),
                'ENTIDAD_NO_FAMILIAR': cuadro_inc['ENTIDAD_NO_FAMILIAR'].sum(),
                'CARGO_DECISION': cuadro_inc['CARGO_DECISION'].sum(),
                'REVISAR': cuadro_inc['REVISAR'].sum(),
                'TOTAL': cuadro_inc['TOTAL'].sum()
            }])
            cuadro_inc = pd.concat([cuadro_inc, total_inc], ignore_index=True)
            
            cuadro_inc.to_excel(writer, sheet_name='Inconsistencias', index=False, startrow=2)
            ws2 = writer.sheets['Inconsistencias']
            ws2['A1'] = 'DISTRIBUCIÓN DE INCONSISTENCIAS POR RAMA Y TIPO'
            ws2['A1'].font = Font(bold=True, size=12)
            ws2.column_dimensions['A'].width = 60
        
        # HOJA 3: CASOS REVISIÓN
        cols_revision = ['directorio', 'secuencia_p', 'orden', 'municipio',
                        'p6370', 'p6380', 'p3069', 'g_p6390s2',
                        'tipo_revision', 'pos_corregida', 'observacion']
        cols_disponibles = [c for c in cols_revision if c in df_fam.columns]
        casos_rev = df_fam[df_fam['tipo_revision'] > 0][cols_disponibles].copy()
        casos_rev.to_excel(writer, sheet_name='Casos_Revision', index=False)
        
        # HOJA 4: TODOS LOS CASOS
        df_fam.to_excel(writer, sheet_name='Casos_Completo', index=False)
    
    output.seek(0)
    return output


def generar_excel_otro(df_otro):
    """Genera Excel para 'Otro, ¿cuál?'."""
    if len(df_otro) == 0:
        return None
    
    # Aplicar clasificación
    clasificaciones = df_otro.apply(clasificar_otro_cual, axis=1, result_type='expand')
    df_otro = df_otro.copy()
    df_otro['tipo_revision'] = clasificaciones['tipo_revision']
    df_otro['pos_corregida'] = clasificaciones['pos_corregida']
    df_otro['observacion'] = clasificaciones['observacion']
    
    # Crear resumen
    resumen = df_otro.groupby('g_p6390s2').agg(
        Casos=('directorio', 'count'),
        Cambiar=('tipo_revision', lambda x: (x.isin([1, 2])).sum()),
        Detallar=('tipo_revision', lambda x: (x == 3).sum()),
        Revisar=('tipo_revision', lambda x: (x == 4).sum())
    ).reset_index()
    resumen = resumen.rename(columns={'g_p6390s2': 'RAMA DE ACTIVIDAD ECONÓMICA'})
    
    # Ordenar y agregar total
    todas_ramas = pd.DataFrame({'RAMA DE ACTIVIDAD ECONÓMICA': ORDEN_RAMAS})
    resumen = todas_ramas.merge(resumen, on='RAMA DE ACTIVIDAD ECONÓMICA', how='left').fillna(0)
    for col in ['Casos', 'Cambiar', 'Detallar', 'Revisar']:
        resumen[col] = resumen[col].astype(int)
    
    total = pd.DataFrame([{
        'RAMA DE ACTIVIDAD ECONÓMICA': 'TOTAL',
        'Casos': resumen['Casos'].sum(),
        'Cambiar': resumen['Cambiar'].sum(),
        'Detallar': resumen['Detallar'].sum(),
        'Revisar': resumen['Revisar'].sum()
    }])
    resumen = pd.concat([resumen, total], ignore_index=True)
    
    # Crear Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # HOJA 1: RESUMEN
        resumen.to_excel(writer, sheet_name='Resumen', index=False, startrow=4)
        ws = writer.sheets['Resumen']
        
        ws['A1'] = 'DISTRIBUCIÓN DE "OTRO, ¿CUÁL?" (P6430=8) POR RAMA DE ACTIVIDAD'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = 'Cambiar = A cuenta propia (5) o patrón (4) | Detallar = Caso ambiguo | Revisar = Posiblemente válido'
        ws['A2'].font = Font(italic=True, size=10)
        ws['A3'] = 'SEMÁFORO:'
        ws['B3'] = '🔴 ROJO = Cambiar posición'
        ws['B3'].fill = ROJO
        ws['C3'] = '🟡 AMARILLO = Detallar'
        ws['C3'].fill = AMARILLO
        ws['D3'] = '🔵 AZUL = Revisar'
        ws['D3'].fill = AZUL
        ws['E3'] = '🟢 VERDE = OK'
        ws['E3'].fill = VERDE
        
        # Aplicar colores
        for i, row in resumen.iterrows():
            fila_excel = i + 6
            if row['RAMA DE ACTIVIDAD ECONÓMICA'] != 'TOTAL':
                if row['Cambiar'] > 0:
                    color = ROJO
                elif row['Detallar'] > 0:
                    color = AMARILLO
                elif row['Revisar'] > 0:
                    color = AZUL
                else:
                    color = VERDE
                for col in range(1, 6):
                    ws.cell(row=fila_excel, column=col).fill = color
        
        ws.column_dimensions['A'].width = 70
        
        # HOJA 2: CUADRO INCONSISTENCIAS
        inconsistencias = df_otro[df_otro['tipo_revision'] > 0].copy()
        if len(inconsistencias) > 0:
            def categorizar(row):
                if row['tipo_revision'] == 1:
                    return 'CUENTA_PROPIA'
                elif row['tipo_revision'] == 2:
                    return 'PATRON'
                elif row['tipo_revision'] == 3:
                    return 'DETALLAR'
                else:
                    return 'REVISAR'
            
            inconsistencias['categoria'] = inconsistencias.apply(categorizar, axis=1)
            cuadro_inc = inconsistencias.groupby(['g_p6390s2', 'categoria']).size().unstack(fill_value=0)
            cuadro_inc = cuadro_inc.reset_index().rename(columns={'g_p6390s2': 'RAMA'})
            
            for col in ['CUENTA_PROPIA', 'PATRON', 'DETALLAR', 'REVISAR']:
                if col not in cuadro_inc.columns:
                    cuadro_inc[col] = 0
            
            cols_orden = ['RAMA', 'CUENTA_PROPIA', 'PATRON', 'DETALLAR', 'REVISAR']
            cuadro_inc = cuadro_inc[cols_orden]
            cuadro_inc['TOTAL'] = cuadro_inc[['CUENTA_PROPIA', 'PATRON', 'DETALLAR', 'REVISAR']].sum(axis=1)
            
            total_inc = pd.DataFrame([{
                'RAMA': 'TOTAL',
                'CUENTA_PROPIA': cuadro_inc['CUENTA_PROPIA'].sum(),
                'PATRON': cuadro_inc['PATRON'].sum(),
                'DETALLAR': cuadro_inc['DETALLAR'].sum(),
                'REVISAR': cuadro_inc['REVISAR'].sum(),
                'TOTAL': cuadro_inc['TOTAL'].sum()
            }])
            cuadro_inc = pd.concat([cuadro_inc, total_inc], ignore_index=True)
            
            cuadro_inc.to_excel(writer, sheet_name='Inconsistencias', index=False, startrow=2)
            ws2 = writer.sheets['Inconsistencias']
            ws2['A1'] = 'DISTRIBUCIÓN DE INCONSISTENCIAS POR RAMA Y TIPO'
            ws2['A1'].font = Font(bold=True, size=12)
            ws2.column_dimensions['A'].width = 60
        
        # HOJA 3: CASOS REVISIÓN
        cols_revision = ['directorio', 'secuencia_p', 'orden', 'municipio',
                        'p6370', 'p6380', 'p6430s1', 'p3069', 'g_p6390s2',
                        'tipo_revision', 'pos_corregida', 'observacion']
        cols_disponibles = [c for c in cols_revision if c in df_otro.columns]
        casos_rev = df_otro[df_otro['tipo_revision'] > 0][cols_disponibles].copy()
        casos_rev.to_excel(writer, sheet_name='Casos_Revision', index=False)
        
        # HOJA 4: TODOS LOS CASOS
        df_otro.to_excel(writer, sheet_name='Casos_Completo', index=False)
    
    output.seek(0)
    return output


# =============================================================================
# INTERFAZ STREAMLIT
# =============================================================================

st.title("📊 Revisión de Ocupados - GEIH")
st.markdown("Genera automáticamente los archivos de validación por posición ocupacional")
st.markdown("*Versión corregida con lógica completa de los notebooks originales*")

st.divider()

# Subir archivo
uploaded_file = st.file_uploader(
    "📁 Sube el archivo de revisión de ocupados",
    type=['xlsx', 'xls'],
    help="Archivo Excel con la base de ocupados para revisión"
)

if uploaded_file:
    with st.spinner("Cargando archivo..."):
        try:
            df = pd.read_excel(uploaded_file)
            st.success(f"✅ Archivo cargado: {len(df):,} registros | {len(df.columns)} columnas")
        except Exception as e:
            st.error(f"Error al cargar el archivo: {e}")
            st.stop()
    
    # Mostrar resumen
    st.subheader("📈 Resumen de casos por posición ocupacional")
    col1, col2, col3, col4 = st.columns(4)
    
    n_gobierno = len(df[df['p6430'] == 2]) if 'p6430' in df.columns else 0
    n_particular = len(df[df['p6430'] == 1]) if 'p6430' in df.columns else 0
    n_familiar = len(df[df['p6430'] == 6]) if 'p6430' in df.columns else 0
    n_otro = len(df[df['p6430'] == 8]) if 'p6430' in df.columns else 0
    
    col1.metric("🏛️ Emp. Gobierno (2)", f"{n_gobierno:,}")
    col2.metric("🏢 Emp. Particular (1)", f"{n_particular:,}")
    col3.metric("👨‍👩‍👧 Trab. Familiar (6)", f"{n_familiar:,}")
    col4.metric("❓ Otro, ¿cuál? (8)", f"{n_otro:,}")
    
    st.divider()
    
    # Opciones de generación
    st.subheader("⚙️ Opciones de generación")
    
    col_opts = st.columns(4)
    gen_gobierno = col_opts[0].checkbox("Empleados Gobierno", value=n_gobierno > 0, disabled=n_gobierno == 0)
    gen_particular = col_opts[1].checkbox("Empleados Particular", value=n_particular > 0, disabled=n_particular == 0)
    gen_familiar = col_opts[2].checkbox("Trabajador Familiar", value=n_familiar > 0, disabled=n_familiar == 0)
    gen_otro = col_opts[3].checkbox("Otro, ¿cuál?", value=n_otro > 0, disabled=n_otro == 0)
    
    st.divider()
    
    # Botón para generar
    if st.button("🚀 Generar archivos de revisión", type="primary", use_container_width=True):
        
        fecha = datetime.now().strftime('%Y%m%d')
        archivos_generados = []
        
        with st.spinner("Procesando archivos..."):
            progress = st.progress(0)
            
            # Empleados del gobierno
            if gen_gobierno and n_gobierno > 0:
                progress.progress(10, "Procesando Empleados del Gobierno...")
                df_gob = df[df['p6430'] == 2].copy()
                excel_gob = generar_excel_gobierno(df_gob)
                if excel_gob:
                    archivos_generados.append(('gobierno', excel_gob, f"rev_empleados_gobierno_{fecha}.xlsx"))
            
            # Empleados particulares
            if gen_particular and n_particular > 0:
                progress.progress(35, "Procesando Empleados Particulares...")
                df_part = df[df['p6430'] == 1].copy()
                excel_part = generar_excel_particular(df_part)
                if excel_part:
                    archivos_generados.append(('particular', excel_part, f"rev_emp_particular_{fecha}.xlsx"))
            
            # Trabajador familiar
            if gen_familiar and n_familiar > 0:
                progress.progress(60, "Procesando Trabajadores Familiares...")
                df_fam = df[df['p6430'] == 6].copy()
                excel_fam = generar_excel_familiar(df_fam)
                if excel_fam:
                    archivos_generados.append(('familiar', excel_fam, f"rev_trabajador_familiar_{fecha}.xlsx"))
            
            # Otro, ¿cuál?
            if gen_otro and n_otro > 0:
                progress.progress(85, "Procesando 'Otro, ¿cuál?'...")
                df_otro = df[df['p6430'] == 8].copy()
                excel_otro = generar_excel_otro(df_otro)
                if excel_otro:
                    archivos_generados.append(('otro', excel_otro, f"rev_otro_cual_{fecha}.xlsx"))
            
            progress.progress(100, "¡Completado!")
        
        if archivos_generados:
            st.success(f"✅ Se generaron {len(archivos_generados)} archivo(s) de revisión")
            
            # Mostrar descargas
            st.subheader("📥 Descargar archivos")
            
            cols = st.columns(len(archivos_generados))
            
            iconos = {'gobierno': '🏛️', 'particular': '🏢', 'familiar': '👨‍👩‍👧', 'otro': '❓'}
            nombres = {'gobierno': 'Emp. Gobierno', 'particular': 'Emp. Particular', 
                       'familiar': 'Trab. Familiar', 'otro': 'Otro, ¿cuál?'}
            asignados = {'gobierno': 'Carolina', 'particular': 'Paula', 
                         'familiar': 'Jeannette', 'otro': 'Jeannette'}
            
            for i, (tipo, excel, filename) in enumerate(archivos_generados):
                with cols[i]:
                    st.download_button(
                        label=f"{iconos.get(tipo, '📄')} {nombres.get(tipo, tipo)}",
                        data=excel,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.caption(f"📌 {asignados.get(tipo, '')}")
        else:
            st.warning("⚠️ No se generaron archivos. Verifica las opciones seleccionadas.")

else:
    st.info("👆 Sube un archivo Excel para comenzar")
    
    with st.expander("ℹ️ ¿Cómo funciona?"):
        st.markdown("""
        ### Proceso de revisión
        1. **Sube** el archivo de revisión de ocupados (Excel con base filtrada)
        2. **Revisa** el resumen de casos por posición ocupacional
        3. **Selecciona** qué posiciones quieres generar
        4. **Genera** los archivos de validación
        5. **Descarga** cada archivo y distribúyelo al equipo
        
        ### Estructura de los archivos generados
        Cada archivo Excel contiene:
        - **Resumen** → Distribución por rama con semáforo de colores
        - **Inconsistencias** → Cuadro resumen de casos a revisar (con TOTAL)
        - **Casos_Revision** → Columnas acotadas para revisar rápido
        - **Casos_Completo** → TODAS las columnas para enviar a campo
        
        ### Asignación de archivos
        - `rev_empleados_gobierno_FECHA.xlsx` → **Carolina**
        - `rev_emp_particular_FECHA.xlsx` → **Paula**
        - `rev_trabajador_familiar_FECHA.xlsx` → **Jeannette**
        - `rev_otro_cual_FECHA.xlsx` → **Jeannette**
        
        ### Semáforo de colores
        - 🔴 **ROJO** = Cambiar posición (inconsistencia clara)
        - 🟡 **AMARILLO** = Detallar/Cambiar rama (requiere verificación)
        - 🔵 **AZUL** = Revisar (caso ambiguo)
        - 🟢 **VERDE** = OK (sin problemas detectados)
        """)
    
    with st.expander("📋 Diccionarios de validación"):
        st.markdown("""
        ### Empleados del Gobierno (P6430=2)
        - Detecta ramas prohibidas donde no debe haber empleados del gobierno
        - Identifica empresas con régimen laboral privado (Ecopetrol)
        - Detecta empresas mixtas donde solo directivos pueden ser pos 2
        - Detecta contratistas que deberían ser cuenta propia
        
        ### Empleados Particulares (P6430=1)
        - Detecta posibles empleados del gobierno (universidades públicas, entidades estatales)
        - Detecta posibles empleados domésticos
        - Detecta posibles jornaleros en agricultura
        
        ### Trabajador Familiar (P6430=6)
        - Detecta casos que trabajan solos (inconsistencia)
        - Detecta entidades no familiares (iglesias, empresas formales)
        - Detecta cargos de decisión (dueño, socio → cuenta propia)
        
        ### Otro, ¿cuál? (P6430=8)
        - Detecta contratistas → cuenta propia
        - Detecta socios/dueños → patrón o cuenta propia
        - Valida casos correctos de "Otro" (subcontratados, madres comunitarias)
        """)

# Footer
st.divider()
st.caption("DANE • DIMPE • Equipo de Validación GEIH")
st.caption("Versión 2.0 - Diciembre 2024")
