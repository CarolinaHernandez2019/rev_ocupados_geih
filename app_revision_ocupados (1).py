"""
Aplicación Streamlit para la Revisión de Ocupados - GEIH
Genera automáticamente los archivos de validación por posición ocupacional
"""

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO
from datetime import datetime

# =============================================================================
# CONFIGURACIÓN DE LA PÁGINA
# =============================================================================
st.set_page_config(
    page_title="Revisión Ocupados GEIH",
    page_icon="📊",
    layout="wide"
)

# =============================================================================
# ESTILOS Y CONSTANTES
# =============================================================================
ROJO = PatternFill('solid', fgColor='FF6B6B')
AMARILLO = PatternFill('solid', fgColor='FFE066')
VERDE = PatternFill('solid', fgColor='8FD14F')
AZUL = PatternFill('solid', fgColor='87CEEB')

ORDEN_RAMAS = [
    'No informa',
    'Agricultura, ganadería, caza, silvicultura y pesca',
    'Explotación de Minas y Canteras',
    'Comercio y reparación de vehículos',
    'Alojamiento y servicios de comida',
    'Industria manufacturera',
    'Suministro de agua y gestión de desechos',
    'Construcción',
    'Transporte y almacenamiento',
    'Información y comunicaciones',
    'Actividades financieras y de seguros',
    'Actividades Inmobiliarias',
    'Actividades profesionales, científicas, técnicas y servicios administrativos',
    'Administración pública y defensa, educación y atención de la salud',
    'Actividades artísticas, entretenimiento, recreación y otras actividades de servicios'
]

# =============================================================================
# DICCIONARIOS DE VALIDACIÓN
# =============================================================================

# Empleados del gobierno
RAMAS_PROHIBIDAS_GOBIERNO = [
    'Agricultura, ganadería, caza, silvicultura y pesca',
    'Explotación de Minas y Canteras',
    'Comercio y reparación de vehículos',
    'Alojamiento y servicios de comida',
    'Construcción'
]

ENTIDADES_GOBIERNO = [
    'ALCALDIA', 'GOBERNACION', 'MINISTERIO', 'SECRETARIA', 'FISCALIA',
    'PROCURADURIA', 'CONTRALORIA', 'DEFENSORIA', 'REGISTRADURIA',
    'POLICIA', 'EJERCITO', 'ARMADA', 'FUERZA AEREA', 'INPEC',
    'DIAN', 'DANE', 'ICBF', 'SENA', 'ICETEX', 'COLPENSIONES',
    'HOSPITAL', 'E.S.E', 'ESE ', 'CLINICA DEL ESTADO',
    'UNIVERSIDAD NACIONAL', 'UNIVERSIDAD DISTRITAL', 'UNIVERSIDAD PEDAGOGICA',
    'COLEGIO DISTRITAL', 'INSTITUCION EDUCATIVA DISTRITAL', 'I.E.D',
    'PERSONERIA', 'CONCEJO', 'ASAMBLEA', 'CONGRESO', 'SENADO', 'CAMARA'
]

PALABRAS_CONTRATISTA = [
    'CONTRATISTA', 'PRESTACION DE SERVICIOS', 'PRESTACIÓN DE SERVICIOS',
    'OPS', 'ORDEN DE PRESTACION', 'CONTRATO DE PRESTACION'
]

# Trabajador familiar
ENTIDADES_NO_FAMILIARES = [
    'IGLESIA', 'PARROQUIA', 'FUNDACION', 'CORPORACION', 'COOPERATIVA',
    'S.A.S', 'SAS', 'S.A', 'LTDA', 'LIMITADA', 'E.S.P', 'ESP',
    'BANCO', 'ALMACEN', 'SUPERMERCADO', 'EXITO', 'JUMBO', 'CARULLA'
]

CARGOS_DECISION = [
    'DUEÑO', 'DUEÑA', 'PROPIETARIO', 'PROPIETARIA', 'GERENTE',
    'ADMINISTRADOR', 'ADMINISTRADORA', 'SOCIO', 'SOCIA', 'PATRON'
]

# Otro, ¿cuál?
PALABRAS_CUENTA_PROPIA = [
    'CONTRATISTA', 'PRESTACION DE SERVICIOS', 'PRESTACIÓN DE SERVICIOS',
    'INDEPENDIENTE', 'FREELANCE', 'POR SU CUENTA'
]

PALABRAS_PATRON = ['SOCIO', 'SOCIA', 'DUEÑO', 'DUEÑA', 'PROPIETARIO', 'ACCIONISTA']

PALABRAS_OTRO_VALIDO = [
    'SUBCONTRATADO', 'SUBCONTRATADA', 'CONTRATADO POR UN ASALARIADO',
    'MADRE COMUNITARIA', 'AYUDANTE DE MADRE', 'OTRO PAIS', 'OTRO PAÍS',
    'HIJO DEL MAYORDOMO', 'HIJA DEL MAYORDOMO'
]

# =============================================================================
# FUNCIONES DE CLASIFICACIÓN
# =============================================================================

def clasificar_empleado_gobierno(row):
    """Clasifica empleados del gobierno (P6430=2)"""
    rama = str(row.get('g_p6390s2', '')).upper() if pd.notna(row.get('g_p6390s2')) else ''
    empresa = str(row.get('p6380', '')).upper() if pd.notna(row.get('p6380')) else ''
    oficio = str(row.get('p6370', '')).upper() if pd.notna(row.get('p6370')) else ''
    
    resultado = {'tipo_revision': 0, 'pos_corregida': None, 'observacion': ''}
    
    # Rama prohibida
    if any(r.upper() in rama for r in RAMAS_PROHIBIDAS_GOBIERNO):
        if any(e in empresa for e in ENTIDADES_GOBIERNO):
            resultado['tipo_revision'] = 4
            resultado['observacion'] = 'REVISAR: Rama atípica pero entidad gobierno'
        else:
            resultado['tipo_revision'] = 1
            resultado['pos_corregida'] = 1
            resultado['observacion'] = 'CAMBIAR: Rama prohibida para gobierno'
        return resultado
    
    # Contratista
    if any(c in oficio or c in empresa for c in PALABRAS_CONTRATISTA):
        resultado['tipo_revision'] = 2
        resultado['pos_corregida'] = 5
        resultado['observacion'] = 'CAMBIAR: Contratista → Cuenta propia'
        return resultado
    
    # Verificar si es entidad gobierno
    if any(e in empresa for e in ENTIDADES_GOBIERNO):
        resultado['tipo_revision'] = 0
        resultado['observacion'] = 'OK: Entidad gobierno confirmada'
    else:
        resultado['tipo_revision'] = 4
        resultado['observacion'] = 'REVISAR: Verificar si es entidad pública'
    
    return resultado


def clasificar_trabajador_familiar(row):
    """Clasifica trabajador familiar sin remuneración (P6430=6)"""
    empresa = str(row.get('p6380', '')).upper() if pd.notna(row.get('p6380')) else ''
    oficio = str(row.get('p6370', '')).upper() if pd.notna(row.get('p6370')) else ''
    p3069 = row.get('p3069', None)
    
    resultado = {'tipo_revision': 0, 'pos_corregida': None, 'observacion': ''}
    
    # Trabaja solo
    try:
        if pd.notna(p3069) and int(p3069) == 1:
            resultado['tipo_revision'] = 1
            resultado['observacion'] = 'DETALLAR: Trabaja solo (P3069=1)'
            return resultado
    except:
        pass
    
    # Entidad no familiar
    if any(e in empresa for e in ENTIDADES_NO_FAMILIARES):
        resultado['tipo_revision'] = 2
        resultado['observacion'] = 'DETALLAR: No parece empresa familiar'
        return resultado
    
    # Cargo de decisión
    if any(c in oficio for c in CARGOS_DECISION):
        resultado['tipo_revision'] = 3
        resultado['pos_corregida'] = 5
        resultado['observacion'] = 'DETALLAR: Cargo decisión → posible cuenta propia'
        return resultado
    
    resultado['observacion'] = 'OK'
    return resultado


def clasificar_otro_cual(row):
    """Clasifica 'Otro, ¿cuál?' (P6430=8)"""
    oficio = str(row.get('p6370', '')).upper() if pd.notna(row.get('p6370')) else ''
    otro_cual = str(row.get('p6430s1', '')).upper() if pd.notna(row.get('p6430s1')) else ''
    p3069 = row.get('p3069', None)
    
    texto = f"{oficio} {otro_cual}"
    resultado = {'tipo_revision': 0, 'pos_corregida': None, 'observacion': ''}
    
    # Cuenta propia
    if any(p in texto for p in PALABRAS_CUENTA_PROPIA):
        resultado['tipo_revision'] = 1
        resultado['pos_corregida'] = 5
        resultado['observacion'] = 'CAMBIAR: Contratista → Cuenta propia (P6430=5)'
        return resultado
    
    # Patrón
    if any(p in texto for p in PALABRAS_PATRON):
        tiene_empleados = False
        try:
            if pd.notna(p3069) and int(p3069) > 1:
                tiene_empleados = True
        except:
            pass
        
        if tiene_empleados:
            resultado['tipo_revision'] = 2
            resultado['pos_corregida'] = 4
            resultado['observacion'] = 'CAMBIAR: Socio/Dueño con empleados → Patrón (P6430=4)'
        else:
            resultado['tipo_revision'] = 1
            resultado['pos_corregida'] = 5
            resultado['observacion'] = 'CAMBIAR: Socio/Dueño sin empleados → Cuenta propia'
        return resultado
    
    # Otro válido
    if any(v in texto for v in PALABRAS_OTRO_VALIDO):
        resultado['tipo_revision'] = 0
        resultado['observacion'] = 'OK: Caso válido de Otro'
        return resultado
    
    # Sin clasificar
    if len(otro_cual.strip()) > 3:
        resultado['tipo_revision'] = 3
        resultado['observacion'] = 'DETALLAR: Verificar descripción'
    else:
        resultado['tipo_revision'] = 3
        resultado['observacion'] = 'DETALLAR: Sin descripción clara'
    
    return resultado


def clasificar_empleado_particular(row):
    """Clasifica empleado particular (P6430=1) - detecta doméstico y jornalero"""
    rama = str(row.get('g_p6390s2', '')).upper() if pd.notna(row.get('g_p6390s2')) else ''
    empresa = str(row.get('p6380', '')).upper() if pd.notna(row.get('p6380')) else ''
    oficio = str(row.get('p6370', '')).upper() if pd.notna(row.get('p6370')) else ''
    
    resultado = {'tipo_revision': 0, 'pos_corregida': None, 'observacion': ''}
    
    # Posible empleado gobierno
    if any(e in empresa for e in ENTIDADES_GOBIERNO):
        resultado['tipo_revision'] = 1
        resultado['pos_corregida'] = 2
        resultado['observacion'] = 'REVISAR: Posible empleado gobierno'
        return resultado
    
    # Posible doméstico
    palabras_domestico = ['EMPLEADA DOMESTICA', 'SERVICIO DOMESTICO', 'HOGAR', 'ASEO EN CASA']
    if any(d in oficio or d in empresa for d in palabras_domestico):
        resultado['tipo_revision'] = 2
        resultado['pos_corregida'] = 3
        resultado['observacion'] = 'REVISAR: Posible empleado doméstico'
        return resultado
    
    # Posible jornalero
    if 'AGRICULTURA' in rama:
        palabras_jornalero = ['JORNALERO', 'PEON', 'COSECHA', 'SIEMBRA', 'ORDEÑO']
        if any(j in oficio for j in palabras_jornalero):
            resultado['tipo_revision'] = 3
            resultado['pos_corregida'] = 7
            resultado['observacion'] = 'REVISAR: Posible jornalero'
            return resultado
    
    resultado['observacion'] = 'OK'
    return resultado


# =============================================================================
# FUNCIÓN PARA GENERAR EXCEL
# =============================================================================

def generar_excel(df_filtrado, titulo, clasificar_func):
    """Genera el archivo Excel con formato para una posición ocupacional"""
    
    if len(df_filtrado) == 0:
        return None
    
    # Aplicar clasificación
    resultados = df_filtrado.apply(clasificar_func, axis=1, result_type='expand')
    df_filtrado = df_filtrado.copy()
    df_filtrado['tipo_revision'] = resultados['tipo_revision']
    df_filtrado['pos_corregida'] = resultados['pos_corregida']
    df_filtrado['observacion'] = resultados['observacion']
    
    # Crear resumen por rama
    resumen = df_filtrado.groupby('g_p6390s2').agg(
        Casos=('tipo_revision', 'count'),
        Cambiar=('tipo_revision', lambda x: (x.isin([1, 2])).sum()),
        Detallar=('tipo_revision', lambda x: (x == 3).sum()),
        Revisar=('tipo_revision', lambda x: (x == 4).sum())
    ).reset_index()
    resumen.columns = ['RAMA DE ACTIVIDAD ECONÓMICA', 'Casos', 'Cambiar', 'Detallar', 'Revisar']
    
    # Agregar total
    total = pd.DataFrame([{
        'RAMA DE ACTIVIDAD ECONÓMICA': 'TOTAL',
        'Casos': resumen['Casos'].sum(),
        'Cambiar': resumen['Cambiar'].sum(),
        'Detallar': resumen['Detallar'].sum(),
        'Revisar': resumen['Revisar'].sum()
    }])
    resumen = pd.concat([resumen, total], ignore_index=True)
    
    # Crear Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Hoja 1: Resumen
        resumen.to_excel(writer, sheet_name='Resumen', index=False, startrow=3)
        ws = writer.sheets['Resumen']
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = 'SEMÁFORO: 🔴 Cambiar | 🟡 Detallar | 🔵 Revisar | 🟢 OK'
        ws['A2'].font = Font(italic=True, size=10)
        
        # Aplicar colores
        for i, row in resumen.iterrows():
            fila = i + 5
            if row['RAMA DE ACTIVIDAD ECONÓMICA'] != 'TOTAL':
                if row['Cambiar'] > 0:
                    color = ROJO
                elif row['Detallar'] > 0:
                    color = AMARILLO
                elif row['Revisar'] > 0:
                    color = AZUL
                else:
                    color = VERDE
                for col in range(1, 6):
                    ws.cell(row=fila, column=col).fill = color
        
        ws.column_dimensions['A'].width = 60
        
        # Hoja 2: Casos para revisión
        cols_revision = ['directorio', 'secuencia_p', 'orden', 'p6370', 'p6380', 
                        'g_p6390s2', 'p6430', 'tipo_revision', 'pos_corregida', 'observacion']
        cols_disponibles = [c for c in cols_revision if c in df_filtrado.columns]
        df_filtrado[cols_disponibles].to_excel(writer, sheet_name='Casos_Revision', index=False)
        
        # Hoja 3: Todos los casos
        df_filtrado.to_excel(writer, sheet_name='Casos_Completo', index=False)
    
    output.seek(0)
    return output


# =============================================================================
# INTERFAZ STREAMLIT
# =============================================================================

st.title("📊 Revisión de Ocupados - GEIH")
st.markdown("Genera automáticamente los archivos de validación por posición ocupacional")

st.divider()

# Subir archivo
uploaded_file = st.file_uploader(
    "📁 Sube el archivo de revisión de ocupados",
    type=['xlsx', 'xls'],
    help="Archivo Excel con la base de ocupados para revisión"
)

if uploaded_file:
    with st.spinner("Cargando archivo..."):
        try:
            df = pd.read_excel(uploaded_file)
            st.success(f"✅ Archivo cargado: {len(df):,} registros")
        except Exception as e:
            st.error(f"Error al cargar el archivo: {e}")
            st.stop()
    
    # Mostrar resumen
    col1, col2, col3, col4 = st.columns(4)
    
    n_gobierno = len(df[df['p6430'] == 2]) if 'p6430' in df.columns else 0
    n_particular = len(df[df['p6430'] == 1]) if 'p6430' in df.columns else 0
    n_familiar = len(df[df['p6430'] == 6]) if 'p6430' in df.columns else 0
    n_otro = len(df[df['p6430'] == 8]) if 'p6430' in df.columns else 0
    
    col1.metric("Emp. Gobierno (2)", f"{n_gobierno:,}")
    col2.metric("Emp. Particular (1)", f"{n_particular:,}")
    col3.metric("Trab. Familiar (6)", f"{n_familiar:,}")
    col4.metric("Otro, ¿cuál? (8)", f"{n_otro:,}")
    
    st.divider()
    
    # Botón para generar
    if st.button("🚀 Generar archivos de revisión", type="primary", use_container_width=True):
        
        fecha = datetime.now().strftime('%Y%m%d')
        archivos_generados = []
        
        with st.spinner("Procesando..."):
            
            # Empleados del gobierno
            if n_gobierno > 0:
                df_gob = df[df['p6430'] == 2].copy()
                excel_gob = generar_excel(df_gob, "REVISIÓN EMPLEADOS DEL GOBIERNO (P6430=2)", clasificar_empleado_gobierno)
                if excel_gob:
                    archivos_generados.append(('gobierno', excel_gob, f"rev_empleados_gobierno_{fecha}.xlsx"))
            
            # Empleados particulares
            if n_particular > 0:
                df_part = df[df['p6430'] == 1].copy()
                excel_part = generar_excel(df_part, "REVISIÓN EMPLEADOS PARTICULARES (P6430=1)", clasificar_empleado_particular)
                if excel_part:
                    archivos_generados.append(('particular', excel_part, f"rev_emp_particular_{fecha}.xlsx"))
            
            # Trabajador familiar
            if n_familiar > 0:
                df_fam = df[df['p6430'] == 6].copy()
                excel_fam = generar_excel(df_fam, "REVISIÓN TRABAJADOR FAMILIAR (P6430=6)", clasificar_trabajador_familiar)
                if excel_fam:
                    archivos_generados.append(('familiar', excel_fam, f"rev_trabajador_familiar_{fecha}.xlsx"))
            
            # Otro, ¿cuál?
            if n_otro > 0:
                df_otro = df[df['p6430'] == 8].copy()
                excel_otro = generar_excel(df_otro, "REVISIÓN OTRO, ¿CUÁL? (P6430=8)", clasificar_otro_cual)
                if excel_otro:
                    archivos_generados.append(('otro', excel_otro, f"rev_otro_cual_{fecha}.xlsx"))
        
        st.success(f"✅ Se generaron {len(archivos_generados)} archivos")
        
        # Mostrar descargas
        st.subheader("📥 Descargar archivos")
        
        cols = st.columns(len(archivos_generados)) if archivos_generados else [st]
        
        iconos = {'gobierno': '🏛️', 'particular': '🏢', 'familiar': '👨‍👩‍👧', 'otro': '❓'}
        nombres = {'gobierno': 'Emp. Gobierno', 'particular': 'Emp. Particular', 
                   'familiar': 'Trab. Familiar', 'otro': 'Otro, ¿cuál?'}
        
        for i, (tipo, excel, filename) in enumerate(archivos_generados):
            with cols[i]:
                st.download_button(
                    label=f"{iconos.get(tipo, '📄')} {nombres.get(tipo, tipo)}",
                    data=excel,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

else:
    st.info("👆 Sube un archivo Excel para comenzar")
    
    with st.expander("ℹ️ ¿Cómo funciona?"):
        st.markdown("""
        1. **Sube** el archivo de revisión de ocupados (Excel)
        2. **Revisa** el resumen de casos por posición
        3. **Genera** los archivos de validación
        4. **Descarga** cada archivo y distribúyelo al equipo
        
        **Archivos generados:**
        - `rev_empleados_gobierno_FECHA.xlsx` → Carolina
        - `rev_trabajador_familiar_FECHA.xlsx` → Jeannette  
        - `rev_otro_cual_FECHA.xlsx` → Jeannette
        - `rev_emp_particular_FECHA.xlsx` → Paula
        """)

# Footer
st.divider()
st.caption("DANE • DIMPE • Equipo de Validación GEIH")
